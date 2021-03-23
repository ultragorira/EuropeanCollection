import os
import xlrd
import json
import openpyxl

base_dir = os.path.dirname(__file__)
input_jsonpath= os.path.join(base_dir,'JSON_Input\JSONDump.xlsx') 
JSON_Folder = os.path.join(base_dir, 'JSON_Output/')

if not os.path.exists(JSON_Folder):
  os.makedirs(JSON_Folder)

wb = openpyxl.load_workbook(input_jsonpath, data_only=True)
sh = wb[wb.sheetnames[0]]
# List to hold dictionaries
json_list = []
fileCount = 0

# Iterate through each row in worksheet and fetch values into dict
for rownum in range(2, sh.max_row+1):
    #data = OrderedDict()
    data = {}
    #row_values = sh.row_values(rownum)
    data['sessionLocalStartTime'] = sh.cell(rownum, 1).value  
    data['audioFileName'] = sh.cell(rownum, 2).value 
    data['collectionId'] = sh.cell(rownum, 3).value 
    data['collectionUuid'] = sh.cell(rownum,4).value 
    data['audioFormat'] = {}
    if (sh.cell(rownum, 5).value == 'Little'):
        endian = "true"
    else:
        endian = "false" 
    data['audioFormat']['bigEndian'] = endian
    data['audioFormat']['channels'] = sh.cell(rownum, 6).value
    data['audioFormat']['encoding'] = sh.cell(rownum, 7).value
    data['audioFormat']['frameRate'] = sh.cell(rownum, 8).value
    data['audioFormat']['frameSize'] = sh.cell(rownum, 9).value
    data['audioFormat']['sampleRate'] = sh.cell(rownum, 10).value
    data['audioFormat']['sampleSizeInBits'] = sh.cell(rownum, 11).value
    data['audioFormat']['audioDurationMillis'] = sh.cell(rownum, 12).value
    data['audioFormat']['format'] = sh.cell(rownum, 13).value
    data['audioFormat']['sourceType'] = sh.cell(rownum, 14).value
    data['sessionInfo'] = {}
    data['sessionInfo']['language_collected'] = sh.cell(rownum, 15).value
    data['sessionInfo']['age_range'] = sh.cell(rownum, 16).value
    data['sessionInfo']['gender'] = sh.cell(rownum, 17).value
    data['sessionInfo']['country_lived'] = sh.cell(rownum, 18).value
    data['sessionInfo']['language_spoken_daily'] = sh.cell(rownum, 19).value
    data['sessionInfo']['primaryMicrophone'] = sh.cell(rownum, 20).value
    data['transcription'] = {}
    data['transcription']['audioDurationMillis'] = sh.cell(rownum, 21).value
    data['transcription']['domain'] = sh.cell(rownum, 22).value
    data['transcription']['id'] = str(sh.cell(rownum, 23).value)
    data['transcription']['phrase'] = sh.cell(rownum, 24).value
    data['transcription']['startEndpoint'] = sh.cell(rownum, 25).value
    data['transcription']['stopEndpoint'] = sh.cell(rownum, 26).value
    data['transcription']['type'] = sh.cell(rownum, 27).value
    data['transcription']['qualityChecks'] = {}
    data['transcription']['wakeword'] = sh.cell(rownum, 29).value
    data['transcription']['intent'] = sh.cell(rownum,30).value
    data['additionalMetadata'] = {}
    data['additionalMetadata']['annotation'] = "null"
    data['additionalMetadata']['dialect'] = sh.cell(rownum,32).value
    data['additionalMetadata']['speakerId'] = sh.cell(rownum, 35).value
    data['additionalMetadata']['phoneModel'] = sh.cell(rownum, 33).value
    data['additionalMetadata']['phoneBrand'] = sh.cell(rownum, 34).value
    data['additionalMetadata']['backgroundNoiseLevelInDb'] = "null"
    data['additionalMetadata']['operatingSystem'] = sh.cell(rownum, 37).value
    data['additionalMetadata']['backgroundNoise'] = sh.cell(rownum, 38).value



    #json_list.append(data)
    # Serialize the list of dicts to JSON
    if data['sessionLocalStartTime'] == "None" or data['sessionLocalStartTime'] == "" or data['sessionLocalStartTime'] == None:
      break
    j = json.dumps(data, indent=4, ensure_ascii=False).encode('utf8')
    # Write to file
    with open(os.path.join(JSON_Folder, data['audioFileName'].replace('wav','') + 'json'), 'wb') as f:
      f.write(j)
      json_list.clear()
    fileCount += 1
print("{} JSON files have been generated!".format(fileCount))
os.startfile(JSON_Folder)